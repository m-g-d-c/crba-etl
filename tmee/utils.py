"""
Utility functions that operate directly on our data-dictionary in Excel using pandas
Typically retrieve queries, useful for our ETL purposes
"""
import requests
import pandas as pd


def get_API_code_address_etc(excel_data_dict):
    """ filters all indicators that are extracted by API
        :param excel_data_dict: path/file to our excel data dictionary in repo
        :return: pandas dataframe (df) with code, address for API requests, data source name and observation footnotes
    """
    # read snapshots table from excel data-dictionary
    snapshot_df = pd.read_excel(excel_data_dict, sheet_name="Snapshot")
    # read sources table from excel data-dictionary
    sources_df = pd.read_excel(excel_data_dict, sheet_name="Source")
    # join snapshot and source based on Source_Id
    # 'left' preserves key order from snapshots
    snap_source_df = snapshot_df.merge(
        sources_df, on="Source_Id", how="left", sort=False
    )
    # read indicators table from excel data-dictionary
    indicators_df = pd.read_excel(excel_data_dict, sheet_name="Indicator")
    # join snap_source and indicators based on Indicator_Id
    snap_source_ind_df = snap_source_df.merge(
        indicators_df, on="Indicator_Id", how="left", sort=False
    )
    # get list of indicator codes, address with API extraction and other info
    logic_API = snap_source_ind_df.Type == "API"
    api_code_addr_etc_df = snap_source_ind_df[logic_API][
        ["Code_y", "Address", "Name_y", "Comments_y"]
    ]
    api_code_addr_etc_df.rename(
        columns={
            "Code_y": "Code",
            "Name_y": "Data_Source",
            "Comments_y": "Obs_Footnote",
        },
        inplace=True,
    )
    # return dataframe (note its df index will correspond to that of snapshots)
    return api_code_addr_etc_df


# function for api request as proposed by Daniele
# errors are printed and don't stop program execution
def api_request(address, params=None, headers=None):
    """
    TODO: Look at the error handerling here
    """
    try:
        response = requests.get(address, params=params, headers=headers)
        # If the response was successful, no Exception will be raised
        response.raise_for_status()
    except requests.exceptions.HTTPError as http_err:
        print(f"HTTP error occurred: {http_err}")
    except Exception as error:
        print(f"Other error occurred: {error}")
    # return response object (check with James note below)
    # for error "no connection adapters found", response will not defined!
    return response


# function from stackoverflow by MaxU
def append_df_to_excel(
    filename,
    df,
    sheet_name="Sheet1",
    startrow=None,
    truncate_sheet=False,
    **to_excel_kwargs,
):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if "engine" in to_excel_kwargs:
        to_excel_kwargs.pop("engine")

    writer = pd.ExcelWriter(filename, engine="openpyxl")

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError
        raise IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    # would this below overwrite startrow defined above? (commented by beto S.)
    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
